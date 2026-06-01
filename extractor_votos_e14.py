"""
=============================================================================
SISTEMA AUTOMATIZADO DE EXTRACCIÓN Y CONSOLIDACIÓN DE DATOS ELECTORALES
Formularios E-14 - Registraduría Nacional del Estado Civil de Colombia
Elecciones Presidenciales 2026
=============================================================================
Versión: 4.0.0
Motor:   Anthropic Claude claude-haiku-4-5 — vía API oficial

Cambios v4.0.0:
  - Migración completa de Ollama a Anthropic Claude API.
  - Claude claude-haiku-4-5: máxima precisión en lectura de marcas electorales.
  - Configuración de API Key vía variable de entorno ANTHROPIC_API_KEY.
  - Parser JSON robusto heredado de v3.3.0.
=============================================================================
"""

import os
import sys
import json
import time
import base64
import logging
import traceback
import re
import platform
import unicodedata
from io import BytesIO
from pathlib import Path
from typing import Optional

from dotenv import load_dotenv
load_dotenv()  # Carga automáticamente el archivo .env del directorio del script

import anthropic
import pandas as pd
from PIL import Image
from pdf2image import convert_from_path

# =============================================================================
# CONFIGURACIÓN GLOBAL
# =============================================================================

RUTA_RAIZ         = Path(r"D:\PROYECTOS VSCODE\conteo_votos_2026_presidente_col")
RUTA_CARPETA_E14  = RUTA_RAIZ / "E14"
RUTA_EXCEL_SALIDA = RUTA_RAIZ / "consolidado_votos_2026.xlsx"

RUTA_POPPLER_WINDOWS: Path | None = (
    RUTA_RAIZ / "Release-26.02.0-0" / "poppler-24.02.0" / "Library" / "bin"
)

DPI_RASTERIZACION = 200      # Mayor DPI = mejor lectura de marcas
JPEG_CALIDAD      = 92
MODELO_CLAUDE     = "claude-haiku-4-5"   # Rápido, económico y preciso

MAX_REINTENTOS    = 3
PAUSA_ENTRE_REINTENTOS = 5  # segundos

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(RUTA_RAIZ / "extraccion_votos.log", encoding="utf-8"),
    ],
)
logger = logging.getLogger(__name__)


# =============================================================================
# PROMPT MAESTRO
# =============================================================================

PROMPT_EXTRACCION = """Eres un experto en lectura de actas electorales (Formulario E-14 Colombia).
Tu única tarea es extraer la información visual y devolver EXCLUSIVAMENTE un JSON válido.

REGLAS DE LECTURA DE VOTOS (COLUMNA "VOTACIÓN"):
La casilla de votación tiene EXACTAMENTE 3 posiciones: [centenas] [decenas] [unidades].
Los jurados rellenan las posiciones VACÍAS con símbolos decorativos: asterisco ✱ * , punto •, guión -, cruz ×.
ESTOS SÍMBOLOS DECORATIVOS NO SON NÚMEROS. IGNÓRALOS COMPLETAMENTE.
Solo los dígitos reales (0 1 2 3 4 5 6 7 8 9) cuentan. Concaténalos de izquierda a derecha.

TABLA DE CONVERSIÓN OBLIGATORIA (aprende estos patrones de memoria):
  ✱ ✱ ✱  →  0     (sin dígitos = cero)
  • • •  →  0
  ✱ ✱ 0  →  0
  ✱ ✱ 1  →  1     ← MUY IMPORTANTE: los dos ✱ NO son el número 2, son relleno vacío
  ✱ ✱ 2  →  2     ← los ✱ delante NO se cuentan
  ✱ ✱ 3  →  3
  ✱ ✱ 5  →  5
  ✱ ✱ 9  →  9
  ✱ 1 0  →  10    ← un ✱ + dígito 1 + dígito 0 = 10
  ✱ 1 5  →  15
  ✱ 2 1  →  21
  ✱ 3 5  →  35
  ✱ 4 8  →  48
  ✱ 7 3  →  73
  ✱ 7 5  →  75
  ✱ 8 2  →  82
  ✱ 9 0  →  90
  1 0 0  →  100
  2 1 7  →  217

REGLA UNIVERSAL: elimina todos los símbolos que NO sean dígitos 0-9, luego lee el número resultante.
Si no queda ningún dígito, el valor es 0.
Si la letra O parece un cero redondo, trátala como 0.

TAMBIÉN extrae estas filas especiales al final de la página 2:
  - "VOTOS EN BLANCO"      → numero: null, partido: "ESPECIAL"
  - "VOTOS NULOS"          → numero: null, partido: "ESPECIAL"
  - "VOTOS NO MARCADOS"    → numero: null, partido: "ESPECIAL"

SUMA TOTAL: en la página 2, última fila "SUMA TOTAL (CANDIDATOS + EN BLANCO + NULOS + NO MARCADOS)".
Extrae ese número y ponlo en "suma_total" del encabezado. Si no aparece, usa 0.

PÁGINA 3 (FIRMAS): si solo hay firmas o constancias de jurados, responde con "candidatos": []

FORMATO DE RESPUESTA (solo JSON puro, sin texto adicional, sin bloques markdown):
{
  "encabezado": {
    "departamento": "TEXTO",
    "municipio": "TEXTO",
    "zona": "TEXTO",
    "puesto": "TEXTO",
    "mesa": "TEXTO",
    "suma_total": 0
  },
  "candidatos": [
    {"numero": 1, "candidato": "NOMBRE COMPLETO", "partido": "NOMBRE AGRUPACION", "votos": 0}
  ]
}"""


# =============================================================================
# UTILIDADES
# =============================================================================

def detectar_ruta_poppler() -> Optional[str]:
    if platform.system() != "Windows":
        return None
    if RUTA_POPPLER_WINDOWS is not None:
        bin_path = RUTA_POPPLER_WINDOWS
        if bin_path.exists() and (bin_path / "pdftoppm.exe").exists():
            logger.info(f"✓ Poppler en ruta configurada: {bin_path}")
            return str(bin_path)
    for encontrado in RUTA_RAIZ.rglob("pdftoppm.exe"):
        carpeta = encontrado.parent
        logger.info(f"✓ Poppler encontrado automáticamente: {carpeta}")
        return str(carpeta)
    logger.warning("Poppler no encontrado. Asegúrate de tenerlo instalado.")
    return None

def imagen_pil_a_base64_jpeg(imagen: Image.Image, calidad: int = JPEG_CALIDAD) -> str:
    buffer = BytesIO()
    imagen.convert("RGB").save(buffer, format="JPEG", quality=calidad, optimize=True)
    return base64.standard_b64encode(buffer.getvalue()).decode("utf-8")

def extraer_json_de_respuesta(texto: str) -> Optional[dict]:
    texto = texto.replace("<|im_end|>", "").strip()
    texto = re.sub(r"```(?:json)?", "", texto).strip("`").strip()
    inicio = texto.find("{")
    fin    = texto.rfind("}")
    if inicio != -1 and fin != -1 and fin > inicio:
        fragmento = texto[inicio : fin + 1]
        try:
            return json.loads(fragmento)
        except json.JSONDecodeError as e:
            logger.warning(f"  └─ JSON inválido: {e}")
    logger.warning("  └─ No se encontró JSON válido en la respuesta.")
    return None

def validar_votos(valor) -> int:
    if valor is None:
        return 0
    try:
        valor_str = str(valor).replace(",", "").replace(" ", "").replace(".", "")
        return max(0, int(float(valor_str)))
    except (ValueError, TypeError):
        return 0

def sanitizar_nombre_hoja(nombre: str, max_len: int = 31) -> str:
    nombre_ascii = unicodedata.normalize("NFKD", nombre)
    nombre_ascii = "".join(c for c in nombre_ascii if not unicodedata.combining(c))
    for c in ['\\', '/', '*', '[', ']', ':', '?']:
        nombre_ascii = nombre_ascii.replace(c, "_")
    return nombre_ascii[:max_len]


# =============================================================================
# CLIENTE ANTHROPIC
# =============================================================================

def crear_cliente_anthropic() -> anthropic.Anthropic:
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        logger.error(
            "ERROR: No se encontró la variable de entorno ANTHROPIC_API_KEY.\n"
            "Configúrala así en PowerShell:\n"
            '  $env:ANTHROPIC_API_KEY = "sk-ant-XXXXXX"\n'
            "O agrega la línea anterior a tu perfil de PowerShell."
        )
        sys.exit(1)
    logger.info(f"✓ API Key de Anthropic cargada (modelo: {MODELO_CLAUDE})")
    return anthropic.Anthropic(api_key=api_key)


def analizar_pagina_con_claude(
    cliente: anthropic.Anthropic,
    imagen_pil: Image.Image,
    nombre_ref: str
) -> Optional[dict]:
    logger.info(f"    → Analizando con Claude {MODELO_CLAUDE}: {nombre_ref}")
    imagen_b64 = imagen_pil_a_base64_jpeg(imagen_pil)

    for intento in range(1, MAX_REINTENTOS + 1):
        try:
            respuesta = cliente.messages.create(
                model=MODELO_CLAUDE,
                max_tokens=2048,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "image",
                                "source": {
                                    "type": "base64",
                                    "media_type": "image/jpeg",
                                    "data": imagen_b64,
                                },
                            },
                            {
                                "type": "text",
                                "text": PROMPT_EXTRACCION,
                            },
                        ],
                    }
                ],
            )
            texto = respuesta.content[0].text.strip()

            if not texto:
                logger.warning(f"    └─ Respuesta vacía (intento {intento}/{MAX_REINTENTOS})")
                time.sleep(PAUSA_ENTRE_REINTENTOS)
                continue

            return extraer_json_de_respuesta(texto)

        except anthropic.RateLimitError:
            logger.warning(f"    └─ Rate limit alcanzado. Esperando 30s (intento {intento})...")
            time.sleep(30)
        except anthropic.APIError as e:
            logger.error(f"    └─ Error API Anthropic (intento {intento}): {e}")
            time.sleep(PAUSA_ENTRE_REINTENTOS)
        except Exception as e:
            logger.error(f"    └─ Error inesperado (intento {intento}): {e}")
            time.sleep(PAUSA_ENTRE_REINTENTOS)

    return None


# =============================================================================
# PROCESAMIENTO DE PDF
# =============================================================================

def procesar_pdf(
    cliente: anthropic.Anthropic,
    ruta_pdf: Path,
    nombre_lugar: str,
    ruta_poppler: Optional[str]
) -> tuple[list[dict], int]:
    registros = []
    nombre_pdf = ruta_pdf.name
    suma_total_pdf = 0   # Valor "SUMA TOTAL" leído del acta
    logger.info(f"  ► {nombre_pdf}")

    try:
        logger.info(f"    → Rasterizando a {DPI_RASTERIZACION} DPI...")
        paginas: list[Image.Image] = convert_from_path(
            str(ruta_pdf), dpi=DPI_RASTERIZACION, fmt="ppm", poppler_path=ruta_poppler
        )

        for num_pag, imagen_pil in enumerate(paginas, start=1):
            nombre_ref = f"{ruta_pdf.stem}_p{num_pag:03d}"
            logger.info(f"    → Página {num_pag}/{len(paginas)}")

            datos = analizar_pagina_con_claude(cliente, imagen_pil, nombre_ref)

            if datos is None:
                continue

            encabezado = datos.get("encabezado", {}) or {}
            candidatos = datos.get("candidatos", []) or []

            # Capturar suma_total cuando la IA la detecte (pág 2)
            st = validar_votos(encabezado.get("suma_total", 0))
            if st > 0:
                suma_total_pdf = st

            if not candidatos:
                logger.info(f"    └─ Página {num_pag}: sin candidatos (hoja de firmas). Omitiendo.")
                continue

            candidatos_ordenados = sorted(
                [c for c in candidatos if c.get("numero") is not None],
                key=lambda x: x.get("numero", 999)
            ) + [c for c in candidatos if c.get("numero") is None]

            # Helper para evitar NoneType.upper()
            def s(val, default="NO LEGIBLE"):
                return (val or default).strip().upper()

            agregados = 0
            for entrada in candidatos_ordenados:
                if not entrada.get("candidato") or entrada.get("candidato") == "NO LEGIBLE":
                    continue
                if "NIVELACIÓN" in str(entrada.get("partido") or "").upper():
                    continue

                registros.append({
                    "Lugar":              nombre_lugar,
                    "Departamento":       s(encabezado.get("departamento")),
                    "Municipio":          s(encabezado.get("municipio")),
                    "Zona":               encabezado.get("zona",   "NO LEGIBLE"),
                    "Puesto":             encabezado.get("puesto", "NO LEGIBLE"),
                    "Mesa":               encabezado.get("mesa",   "NO LEGIBLE"),
                    "Candidato":          s(entrada.get("candidato"), "DESCONOCIDO"),
                    "Partido_Agrupacion": s(entrada.get("partido"),   "DESCONOCIDO"),
                    "Votos":              validar_votos(entrada.get("votos", 0)),
                    "Archivo_PDF":        nombre_pdf,
                    "Pagina":             num_pag,
                })
                agregados += 1

            logger.info(f"    └─ Página {num_pag}: {agregados} registro(s) extraído(s)")

    except Exception as e:
        logger.error(f"  └─ Error procesando '{nombre_pdf}': {e}")
        logger.debug(traceback.format_exc())

    return registros, suma_total_pdf


# =============================================================================
# EXPORTACIÓN A EXCEL
# =============================================================================

def _escribir_hoja(hoja, df_lugar, col_names, totales_pdf,
                   col_votos_idx, col_votos_letra,
                   fill_ia, fill_calc, font_bold, font_header):
    """
    Escribe la hoja fila a fila con un cursor para evitar solapamientos.
    Estructura por PDF:
        [cabecera]
        [datos PDF-1]
        [TOTAL ACTA (IA)  — PDF-1]   ← fila azul
        [TOTAL CALCULADO  — PDF-1]   ← fila verde con fórmula SUM
        [datos PDF-2]
        [TOTAL ACTA (IA)  — PDF-2]
        [TOTAL CALCULADO  — PDF-2]
        ...
    """
    from openpyxl.styles import Font, PatternFill, Alignment

    num_cols = len(col_names)

    # ── Cabecera ──
    for c, col_name in enumerate(col_names, start=1):
        cell = hoja.cell(row=1, column=c, value=col_name)
        cell.font = font_header

    fila = 2  # cursor de escritura
    pdfs_en_hoja = list(dict.fromkeys(df_lugar["Archivo_PDF"].tolist()))

    for nombre_pdf in pdfs_en_hoja:
        df_pdf = df_lugar[df_lugar["Archivo_PDF"] == nombre_pdf]
        fila_inicio = fila

        # ── Datos del grupo ──
        for _, row in df_pdf.iterrows():
            for c, col_name in enumerate(col_names, start=1):
                hoja.cell(row=fila, column=c, value=row[col_name])
            fila += 1

        fila_fin = fila - 1

        # ── TOTAL ACTA (IA) ──
        total_ia = totales_pdf.get(nombre_pdf, 0)
        hoja.cell(row=fila, column=col_votos_idx - 2,
                  value=f"TOTAL ACTA (IA) — {nombre_pdf}")
        hoja.cell(row=fila, column=col_votos_idx, value=total_ia)
        for c in range(1, num_cols + 1):
            cell = hoja.cell(row=fila, column=c)
            cell.fill = fill_ia
            cell.font = font_bold
        fila += 1

        # ── TOTAL CALCULADO (fórmula SUM) ──
        formula = f"=SUM({col_votos_letra}{fila_inicio}:{col_votos_letra}{fila_fin})"
        hoja.cell(row=fila, column=col_votos_idx - 2,
                  value=f"TOTAL CALCULADO — {nombre_pdf}")
        hoja.cell(row=fila, column=col_votos_idx, value=formula)
        for c in range(1, num_cols + 1):
            cell = hoja.cell(row=fila, column=c)
            cell.fill = fill_calc
            cell.font = font_bold
        fila += 1


def exportar_a_excel(registros: list[dict], totales_pdf: dict[str, int]) -> None:
    """
    Genera el Excel con:
      - Hoja CONSOLIDADO: todos los registros con totales por PDF al final de cada grupo.
      - Una hoja por lugar: misma estructura con totales intercalados tras cada PDF.
    """
    if not registros:
        logger.warning("No hay registros para exportar.")
        return

    logger.info(f"\nEXPORTANDO {len(registros)} REGISTROS A EXCEL...")

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill

    df = pd.DataFrame(registros)
    col_names     = list(df.columns)
    col_votos_idx = col_names.index("Votos") + 1          # 1-based
    col_votos_letra = get_column_letter(col_votos_idx)

    # Estilos reutilizables
    fill_ia    = PatternFill("solid", fgColor="D9E1F2")   # azul claro
    fill_calc  = PatternFill("solid", fgColor="E2EFDA")   # verde claro
    font_bold  = Font(bold=True)
    font_header = Font(bold=True)

    wb = Workbook()

    # ── Hoja CONSOLIDADO: resumen total por candidato ─────────────────────────
    hoja_cons = wb.active
    hoja_cons.title = "CONSOLIDADO"

    # Agrupar votos por candidato sumando todos los PDFs y lugares
    resumen = (
        df.groupby("Candidato", sort=False)["Votos"]
        .sum()
        .reset_index()
        .sort_values("Votos", ascending=False)
        .reset_index(drop=True)
    )

    # Cabecera
    hoja_cons.cell(row=1, column=1, value="Candidato").font  = font_header
    hoja_cons.cell(row=1, column=2, value="Total Votos").font = font_header

    # Datos fila a fila
    for i, row in resumen.iterrows():
        hoja_cons.cell(row=i + 2, column=1, value=row["Candidato"])
        hoja_cons.cell(row=i + 2, column=2, value=int(row["Votos"]))

    # Fila GRAN TOTAL con fórmula SUM
    fila_gt = len(resumen) + 2
    cell_lbl = hoja_cons.cell(row=fila_gt, column=1, value="TOTAL DE VOTOS E14 (SUMA DE TODOS LOS CANDIDATOS)")
    cell_lbl.font = font_bold
    cell_lbl.fill = PatternFill("solid", fgColor="FFD966")
    cell_tot = hoja_cons.cell(row=fila_gt, column=2, value=f"=SUM(B2:B{fila_gt - 1})")
    cell_tot.font = font_bold
    cell_tot.fill = PatternFill("solid", fgColor="FFD966")

    # Anchos de columna
    max_cand = resumen["Candidato"].astype(str).map(len).max() if not resumen.empty else 20
    hoja_cons.column_dimensions["A"].width = min(max_cand + 2, 50)
    hoja_cons.column_dimensions["B"].width = 14

    logger.info(f"  ✓ Hoja 'CONSOLIDADO': {len(resumen)} candidatos únicos")

    # ── Una hoja por lugar ────────────────────────────────────────────────────
    lugares = list(dict.fromkeys(df["Lugar"].tolist()))
    for lugar in lugares:
        df_lugar = df[df["Lugar"] == lugar].reset_index(drop=True)
        nombre_hoja = sanitizar_nombre_hoja(lugar)
        hoja = wb.create_sheet(title=nombre_hoja)

        _escribir_hoja(hoja, df_lugar, col_names, totales_pdf,
                       col_votos_idx, col_votos_letra,
                       fill_ia, fill_calc, font_bold, font_header)

        # Ajuste de anchos
        for col_idx, col_name in enumerate(col_names, start=1):
            max_len = max(len(str(col_name)),
                          df_lugar[col_name].astype(str).map(len).max() if not df_lugar.empty else 0)
            hoja.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

        pdfs_en_lugar = list(dict.fromkeys(df_lugar["Archivo_PDF"].tolist()))
        logger.info(f"  ✓ Hoja '{nombre_hoja}': {len(df_lugar)} registro(s) "
                    f"+ {len(pdfs_en_lugar)*2} filas de totales")

    wb.save(str(RUTA_EXCEL_SALIDA))
    logger.info(f"✓ Excel guardado en: {RUTA_EXCEL_SALIDA}")


# =============================================================================
# MAIN
# =============================================================================

def main() -> None:
    cliente = crear_cliente_anthropic()
    ruta_poppler = detectar_ruta_poppler()

    subcarpetas = sorted(d for d in RUTA_CARPETA_E14.iterdir() if d.is_dir())
    if not subcarpetas:
        logger.error(f"No se encontraron subcarpetas en {RUTA_CARPETA_E14}")
        sys.exit(1)

    todos_los_registros: list[dict] = []
    totales_pdf: dict[str, int] = {}   # nombre_pdf → suma_total leída del acta

    for carpeta in subcarpetas:
        nombre_lugar = carpeta.name
        pdfs = sorted({p.name.lower(): p for p in carpeta.glob("*") if p.suffix.lower() == ".pdf"}.values())
        if not pdfs:
            continue
        logger.info(f"\n{'='*60}")
        logger.info(f"LUGAR: {nombre_lugar} ({len(pdfs)} PDF(s))")
        logger.info(f"{'='*60}")
        for ruta_pdf in pdfs:
            registros_pdf, suma_total = procesar_pdf(cliente, ruta_pdf, nombre_lugar, ruta_poppler)
            todos_los_registros.extend(registros_pdf)
            totales_pdf[ruta_pdf.name] = suma_total
            if suma_total:
                logger.info(f"    → Suma total leída del acta: {suma_total}")

    exportar_a_excel(todos_los_registros, totales_pdf)
    logger.info("\nPROCESO FINALIZADO.")


if __name__ == "__main__":
    main()
